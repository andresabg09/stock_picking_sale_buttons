from odoo import models, fields
import math
import io


class PurchaseOrderLine(models.Model):
    _inherit = 'purchase.order.line'

    custom_product_image = fields.Binary(
        related='product_id.image_1024',
        string='Imagen',
        readonly=True,
    )
    custom_product_barcode = fields.Char(
        related='product_id.barcode',
        string='Código de Barras',
        readonly=True,
    )


class PurchaseOrder(models.Model):
    _inherit = 'purchase.order'

    custom_price_list = fields.Selection(
        [
            ('pagoda', 'La Pagoda'),
            ('estandar', 'Precios Estándar Shalom'),
            ('distribuidora', 'Distribuidora'),
        ],
        string='Lista de Precios (Correo)',
        help='Se usa para completar el cuerpo del correo de envío al proveedor.',
    )

    def action_merge(self):
        # Identificar la orden más antigua (la que sobrevive) antes del merge
        rfq_to_merge = self.filtered(lambda r: r.state in ['draft', 'sent'])
        oldest_rfq = min(rfq_to_merge, key=lambda r: r.date_order)

        result = super().action_merge()

        # Operar sobre la orden sobreviviente
        if oldest_rfq and oldest_rfq.exists():
            # Redondeo hacia arriba al múltiplo de 5 para productos TINTE NNP
            for line in oldest_rfq.order_line.filtered(
                lambda l: l.display_type not in ['line_section', 'line_note']
                          and 'TINTE NNP' in (l.product_id.name or '').upper()
            ):
                line.product_qty = math.ceil(line.product_qty / 5) * 5

            # Reordenar líneas alfabéticamente
            sequence = 1
            lines = oldest_rfq.order_line.filtered(
                lambda l: l.display_type not in ['line_section', 'line_note']
            ).sorted(key=lambda l: (l.product_id.name or '').lower())
            for line in lines:
                line.sequence = sequence
                sequence += 1

        return result

    def _generate_xlsx_attachment(self, filename=None):
        """Genera el adjunto XLSX (Código de Barras, Código Anclado, Producto,
        Cantidad) para esta orden de compra y devuelve el ir.attachment creado.
        Si no se pasa filename, usa el nombre de la orden."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        order = self[0]

        if not filename:
            safe_name = order.name.replace('/', '_')
            filename = "%s.xlsx" % safe_name

        wb = Workbook()
        ws = wb.active
        ws.title = "Orden de Compra"

        headers = ["Código de Barras", "Código Anclado", "Producto", "Cantidad"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        BarcodeMulti = self.env['product.barcode.multi']

        for line in order.order_line:
            if line.display_type:
                continue

            product = line.product_id
            barcode = product.barcode or ""

            additional = BarcodeMulti.search([('product_id', '=', product.id)])
            codigo_anclado = ", ".join(additional.mapped('name')) if additional else ""

            prod_name = product.display_name or ""

            ws.append([barcode, codigo_anclado, prod_name, line.product_qty])

        column_widths = [20, 25, 45, 12]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'raw': buffer.read(),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': 'purchase.order',
            'res_id': order.id,
        })
        return attachment

    def _get_xlsx_bytes(self):
        """Igual que _generate_xlsx_attachment pero devuelve los bytes del
        XLSX directamente, sin crear un ir.attachment (para uso interno en
        generación de ZIP masivo)."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        order = self[0]

        wb = Workbook()
        ws = wb.active
        ws.title = "Orden de Compra"

        headers = ["Código de Barras", "Código Anclado", "Producto", "Cantidad"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        BarcodeMulti = self.env['product.barcode.multi']

        for line in order.order_line:
            if line.display_type:
                continue

            product = line.product_id
            barcode = product.barcode or ""

            additional = BarcodeMulti.search([('product_id', '=', product.id)])
            codigo_anclado = ", ".join(additional.mapped('name')) if additional else ""

            prod_name = product.display_name or ""

            ws.append([barcode, codigo_anclado, prod_name, line.product_qty])

        column_widths = [20, 25, 45, 12]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def action_download_xlsx(self):
        """Descarga directa del XLSX (botón 'Descargar Orden de Compra')."""
        attachment = self._generate_xlsx_attachment()
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }

    def action_download_individual_xlsx_zip(self):
        """Acción masiva: genera un ZIP con un XLSX por cada orden de compra
        seleccionada, nombrando cada archivo con el cliente, el número de
        orden de compra y el número de orden de venta (documento origen)."""
        import zipfile
        import re

        zip_buffer = io.BytesIO()
        used_names = {}

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for order in self:
                xlsx_bytes = order._get_xlsx_bytes()

                sale_order_names = []
                if order.origin:
                    parts = order.origin.split(',')
                    for part in parts:
                        so_name = part.strip().split('/')[0].strip()
                        if so_name:
                            sale_order_names.append(so_name)

                sale_orders = self.env['sale.order']
                if sale_order_names:
                    sale_orders = self.env['sale.order'].search([('name', 'in', sale_order_names)])

                clientes = sale_orders.mapped('partner_id.name')
                cliente_str = " y ".join(clientes) if clientes else "Sin Cliente"

                so_str = "_".join(sale_order_names) if sale_order_names else "SinOrigen"

                def clean(text):
                    text = re.sub(r'[\\/:*?"<>|]', '', text)
                    return text.strip()

                base_filename = "%s - %s - %s" % (clean(cliente_str), order.name, so_str)
                filename = "%s.xlsx" % base_filename

                if filename in used_names:
                    used_names[filename] += 1
                    filename = "%s (%s).xlsx" % (base_filename, used_names[filename])
                else:
                    used_names[filename] = 0

                zf.writestr(filename, xlsx_bytes)

        zip_buffer.seek(0)

        zip_attachment = self.env['ir.attachment'].create({
            'name': 'Ordenes_Compra_Individuales.zip',
            'type': 'binary',
            'raw': zip_buffer.read(),
            'mimetype': 'application/zip',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % zip_attachment.id,
            'target': 'self',
        }

    def action_send_email_with_xlsx(self):
        """Abre el composer de correo con el mismo formato usado en el envío
        masivo (saludo dinámico, marcadores en negrita, CC fijo sin duplicar
        el destinatario), pero en singular y sin mencionar la orden de venta,
        ya que aquí se envía una sola orden de compra."""
        self.ensure_one()

        xlsx_attachment = self._generate_xlsx_attachment()

        Wizard = self.env['purchase.bulk.email.wizard']
        saludo = Wizard._saludo_por_hora()

        subject = "Pedido de ruta de [ESCRIBIR RUTA AQUÍ] - Orden de compra %s" % self.name

        body = """
<div style="margin: 0px; padding: 0px;">
    <p style="margin: 0px; padding: 0px; font-size: 13px;">
        %s,
        <br/><br/>
        Les adjunto la ruta de <strong>[ESCRIBIR RUTA AQUÍ]</strong> de la orden de compra
        <strong>%s</strong>
        correspondiente a los precios de <strong>[SELECCIONAR LISTA DE PRECIOS]</strong>.
        <br/><br/>
        Así mismo, solicitamos su especial atención a las siguientes especificaciones de despacho:
        <br/><br/>
        <ul>
            <li>Presentación de Productos NNP: Los artículos Aliset NNP de 69gr y Decolorantes se solicitan como unidades individuales.</li>
            <li>En cambio los AER POCKETS se encuentran por <strong>DISPLAYS</strong>.</li>
        </ul>
        <br/>
        Quedo a su disposición ante cualquier consulta adicional.
        <br/><br/>
        ¿Podría confirmar que recibió esta orden?
        <br/><br/>
        Atentamente,<br/>
        Andrés Gutiérrez<br/>
        Asistente<br/>
        Shalom Panamá.
        <br/><br/>
    </p>
</div>
""" % (saludo, self.name)

        email_to = self.partner_id.email or ''
        cc_emails_str = Wizard._build_cc_for_partner(email_to)
        cc_emails = [e.strip() for e in cc_emails_str.split(',') if e.strip()]

        cc_partner_ids = []
        Partner = self.env['res.partner']
        for email in cc_emails:
            partner = Partner.search([('email', '=ilike', email)], limit=1)
            if not partner:
                partner = Partner.create({'name': email, 'email': email})
            cc_partner_ids.append(partner.id)

        partner_ids = (self.partner_id.ids if self.partner_id else []) + cc_partner_ids

        ctx = {
            'default_model': 'purchase.order',
            'default_res_ids': self.ids,
            'default_use_template': False,
            'default_composition_mode': 'comment',
            'default_subject': subject,
            'default_body': body,
            'default_partner_ids': partner_ids,
            'default_attachment_ids': [(4, xlsx_attachment.id)],
            'mark_so_as_sent': True,
        }
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'mail.compose.message',
            'view_mode': 'form',
            'target': 'new',
            'context': ctx,
        }

    def _xlsx_filename_for_order(self, order):
        """Nombre de archivo: Pedido [N°OrdenVenta/Documento Origen] ([N°OrdenCompra]).xlsx"""
        import re

        sale_order_names = []
        if order.origin:
            parts = order.origin.split(',')
            for part in parts:
                so_name = part.strip().split('/')[0].strip()
                if so_name:
                    sale_order_names.append(so_name)

        so_str = ", ".join(sale_order_names) if sale_order_names else "Sin Origen"

        def clean(text):
            text = re.sub(r'[\\/:*?"<>|]', '', text)
            return text.strip()

        return "Pedido %s (%s).xlsx" % (clean(so_str), clean(order.name))

    def action_send_bulk_email_by_partner(self):
        """Acción masiva: agrupa las órdenes de compra seleccionadas por
        proveedor y abre el wizard 'Enviar Órdenes por Proveedor' para
        procesarlas una por una, cada correo con todos los XLSX de ese
        proveedor adjuntos."""
        if not self:
            return {'type': 'ir.actions.act_window_close'}

        wizard = self.env['purchase.bulk.email.wizard'].create({
            'purchase_order_ids': [(6, 0, self.ids)],
        })
        return wizard.action_open_next_partner()
