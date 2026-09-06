from odoo import models, fields, api

TINTE_NNP_MIN_PRICE = 1.16

DIANKE_EMAIL_TO = "ventasdianke@gmail.com,Dianazuniga@diankegroup.com,kenniarueda@diankegroup.com"
DIANKE_EMAIL_CC = "andres@shalompma.com,luis@shalompma.com,milciades@shalompma.com"

DIANKE_GODREJ_POCKET_UNITS_PER_DISPLAY = 6  # las 6 referencias AMB GODREJ POCKET se manejan por display; Dianke las necesita en unidades

PAYMENT_METHOD_SELECTION = [
    ('efectivo', 'Efectivo'),
    ('tarjeta', 'Tarjeta'),
    ('credito_1_semana', 'Crédito 1 semana'),
    ('credito_2_semanas', 'Crédito 2 semanas'),
    ('transferencia', 'Transferencia'),
    ('yappy', 'Yappy'),
    ('otro', 'Otro'),
]


class SaleOrder(models.Model):
    _inherit = 'sale.order'

    custom_first_product_image = fields.Binary(
        string='Imagen',
        compute='_compute_custom_first_product_image',
        readonly=True,
    )
    custom_payment_method = fields.Selection(
        PAYMENT_METHOD_SELECTION,
        string='Forma de Pago',
        help='Cómo va a pagar el cliente. Se incluye en el reporte que se envía a Dianke.',
    )
    custom_dianke_exported = fields.Boolean(
        string='Enviada a Dianke',
        default=False,
        copy=False,
        help='Se marca automáticamente cuando la orden se incluye en un envío (manual o automático) a Dianke.',
    )
    custom_dianke_exported_date = fields.Datetime(
        string='Fecha envío a Dianke',
        readonly=True,
        copy=False,
    )
    custom_dianke_exported_by = fields.Many2one(
        'res.users',
        string='Enviado a Dianke por',
        readonly=True,
        copy=False,
        help='Usuario que confirmó el envío a Dianke (manual o automático).',
    )

    @api.depends('order_line.product_id')
    def _compute_custom_first_product_image(self):
        for order in self:
            first_line = order.order_line.filtered(
                lambda l: l.product_id and l.product_id.image_128
            )[:1]
            order.custom_first_product_image = (
                first_line.product_id.image_128 if first_line else False
            )

    # ------------------------------------------------------------------
    # Exportación a Dianke
    # ------------------------------------------------------------------

    @staticmethod
    def _dianke_contact_info(partner):
        """Devuelve (nombre_local, ruc, telefono, celular, nombre_contacto,
        direccion) para el reporte de Dianke, a partir de campos de
        res.partner: los estándar de Odoo (vat, phone, mobile) más el campo
        de Studio `x_nombre_contacto` (nombre de la persona, ej. "Julio";
        confirmado por Andrés 2026-09-05 — separado de `name`, que es el
        nombre del local/negocio)."""
        if not partner:
            return ('', '', '', '', '', '')

        company = partner.commercial_partner_id or partner
        nombre_contacto = getattr(partner, 'x_nombre_contacto', '') or ''

        # Se arma solo con los campos de dirección (calle, ciudad, provincia,
        # país) — sin el nombre del cliente, que campos como
        # contact_address_complete traen pegado al inicio.
        partes_direccion = [
            partner.street,
            partner.street2,
            partner.city,
            partner.state_id.name if partner.state_id else '',
            partner.country_id.name if partner.country_id else '',
        ]
        direccion = ", ".join(p for p in partes_direccion if p)

        return (
            partner.name or '',
            partner.vat or company.vat or '',
            partner.phone or company.phone or '',
            partner.mobile or company.mobile or '',
            nombre_contacto,
            direccion,
        )

    def _dianke_route_info(self, partner):
        """(nombre_ruta, orden_en_ruta) del cliente, buscando el
        fsm.location cuyo partner_id es este cliente — mapa de campos
        confirmado por Andrés 2026-09-05. ('', None) si no se encuentra o
        el módulo de rutas no está instalado (no debe tumbar el export)."""
        if not partner:
            return '', None
        try:
            location = self.env['fsm.location'].search([('partner_id', '=', partner.id)], limit=1)
            if not location:
                return '', None
            ruta = location.fsm_route_id.name if location.fsm_route_id else ''
            orden = getattr(location, 'x_orden_ruta', None)
            return ruta, orden
        except Exception:
            return '', None

    @staticmethod
    def _dianke_delivery_date(order_date):
        """4 días hábiles (lunes a viernes, sin contar sábado ni domingo)
        desde la fecha de la orden — regla confirmada por Andrés
        2026-09-05. No contempla feriados, solo fines de semana. No hay
        ningún campo en el sistema que ya calcule esto, así que se computa
        aquí mismo, sin guardar nada nuevo en la orden."""
        if not order_date:
            return None
        from datetime import timedelta
        d = order_date.date() if hasattr(order_date, 'date') else order_date
        dias_habiles = 0
        while dias_habiles < 4:
            d = d + timedelta(days=1)
            if d.weekday() < 5:  # 0=lunes ... 4=viernes (5=sábado, 6=domingo se saltan)
                dias_habiles += 1
        return d

    @staticmethod
    def _dianke_fecha_larga_es(fecha):
        """"Sábado 5 de Septiembre de 2026" — para el asunto del correo a
        Dianke, así se distingue de un vistazo si el envío es de hoy o de
        otro día (ej. si un día se manda solo una parte de una ruta y al
        día siguiente el resto) — pedido de Andrés 2026-09-05. Se escribe
        a mano en vez de usar locale del servidor, que puede no tener
        español instalado."""
        if not fecha:
            return ''
        dias = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
        meses = [
            'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio',
            'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
        ]
        return "%s %d de %s de %d" % (dias[fecha.weekday()], fecha.day, meses[fecha.month - 1], fecha.year)

    @staticmethod
    def _dianke_payment_checkboxes(payment_method):
        """Traduce custom_payment_method a las casillas de la plantilla de
        Dianke (Efectivo/Tarjeta/ACH). Transferencia se marca como ACH (es
        pago electrónico). Yappy y Crédito no tienen casilla propia en su
        plantilla, así que se agrega una 4ta opción "Otro: <forma de pago>"
        para no perder esa información — pedido de Andrés 2026-09-05."""
        efectivo = payment_method == 'efectivo'
        tarjeta = payment_method == 'tarjeta'
        ach = payment_method == 'transferencia'
        otro_label = None
        if payment_method and not (efectivo or tarjeta or ach):
            otro_label = dict(PAYMENT_METHOD_SELECTION).get(payment_method, payment_method)
        return efectivo, tarjeta, ach, otro_label

    def _dianke_order_rows_data(self):
        """Arma, para cada orden de self, un dict con todos los datos ya
        resueltos (cliente, ruta, orden en la ruta, fecha de entrega,
        líneas, etc.)."""
        payment_labels = dict(PAYMENT_METHOD_SELECTION)
        data = []
        for order in self.sorted(key=lambda o: o.name):
            partner = order.partner_id
            local, ruc, telefono, celular, contacto, direccion = self._dianke_contact_info(partner)
            ruta, orden_ruta = self._dianke_route_info(partner)
            data.append({
                'order': order,
                'partner': partner,
                'local': local,
                'ruc': ruc,
                'telefono': telefono,
                'celular': celular,
                'contacto': contacto,
                'direccion': direccion,
                'forma_pago': payment_labels.get(order.custom_payment_method, order.custom_payment_method or ''),
                'fecha': order.date_order.strftime('%d/%m/%Y') if order.date_order else '',
                'ruta': ruta,
                'orden_ruta': orden_ruta,
                'fecha_entrega': self._dianke_delivery_date(order.date_order),
                'lines': order.order_line.filtered(lambda l: not l.display_type),
            })
        return data

    @staticmethod
    def _dianke_group_by_route(rows_data):
        """Agrupa las filas por ruta y, dentro de cada ruta, las ordena por
        la posición del cliente en la ruta (x_orden_ruta), de menor a
        mayor — pedido de Andrés 2026-09-05: un Excel por ruta, con los
        pedidos en el mismo orden en que se van a atender. Sin ruta
        asignada cae en el grupo "Sin Ruta".

        Fix 2026-09-05: x_orden_ruta es un Integer de Odoo — si nunca se
        asignó, su valor por defecto es 0, no None. Antes solo se trataba
        None como "sin asignar", así que todos los clientes con 0 (la
        mayoría, en la práctica) se ordenaban de PRIMEROS en vez de al
        final. Ahora 0 y None se tratan igual (sin asignar → al final)."""
        groups = {}
        group_order = []
        for data in rows_data:
            key = data['ruta'] or 'Sin Ruta'
            if key not in groups:
                groups[key] = []
                group_order.append(key)
            groups[key].append(data)
        for key in groups:
            groups[key].sort(key=lambda d: (not d['orden_ruta'], d['orden_ruta'] or 0))
        return [(key, groups[key]) for key in group_order]

    @staticmethod
    def _dianke_safe_sheet_name(name, used_names):
        """Nombre de pestaña de Excel válido y único: sin :\\/?*[], máximo
        31 caracteres, sin repetirse dentro del mismo archivo."""
        import re
        base = re.sub(r'[:\\/?*\[\]]', '', name or 'Pedido').strip()[:31] or 'Pedido'
        candidate = base
        i = 2
        while candidate in used_names:
            suffix = " (%d)" % i
            candidate = base[:31 - len(suffix)] + suffix
            i += 1
        used_names.add(candidate)
        return candidate

    def _generate_dianke_xlsx_files(self):
        """Genera un XLSX por cada ruta de las órdenes de self (pedido de
        Andrés 2026-09-05), con una pestaña por cliente/pedido, en el mismo
        orden de la ruta. Devuelve una lista de (nombre_archivo, bytes)."""
        from openpyxl import Workbook
        import io
        import re

        rows_data = self._dianke_order_rows_data()
        grouped = self._dianke_group_by_route(rows_data)
        fecha_str = fields.Date.context_today(self).strftime('%d-%m-%Y')

        files = []
        for route_name, group_rows in grouped:
            wb = Workbook()
            used_names = set()
            for i, data in enumerate(group_rows):
                sheet_name = self._dianke_safe_sheet_name(data['local'], used_names)
                ws = wb.active if i == 0 else wb.create_sheet()
                ws.title = sheet_name
                self._fill_dianke_template_sheet(ws, [data])

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            safe_route = re.sub(r'[\\/:*?"<>|]', '', route_name).strip() or 'Sin Ruta'
            filename = "Pedidos Dianke %s %s.xlsx" % (safe_route, fecha_str)
            files.append((filename, buffer.read()))

        return files

    @staticmethod
    def _dianke_embed_partner_photo(ws, partner, anchor_cell, row_for_height, height=90):
        """Incrusta la foto del contacto (si tiene) en anchor_cell, YA
        reducida de peso real con Pillow (no solo redimensionada en
        pantalla) — fix 2026-09-05: antes solo se cambiaba el tamaño
        visual (xl_img.width/height), pero el binario embebido seguía
        siendo la imagen original de Odoo (cientos de KB a varios MB).
        Con muchos clientes en un mismo archivo de ruta, eso hacía que el
        Excel pesara varios MB, se demorara en descargar y que el
        navegador lo bloqueara como "descarga no segura". Ahora se
        comprime a JPEG de baja resolución antes de incrustarla, dejando
        cada foto en unos pocos KB. Devuelve True si se incrustó algo."""
        if not partner or not partner.image_1920:
            return False
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage
        import base64
        import io
        try:
            img_bytes = base64.b64decode(partner.image_1920)
            pil_img = PILImage.open(io.BytesIO(img_bytes))
            pil_img = pil_img.convert('RGB')
            pixel_size = int(height * 2)  # un poco más de resolución que el tamaño mostrado, para que no se vea pixelada
            pil_img.thumbnail((pixel_size, pixel_size))
            resized_buffer = io.BytesIO()
            pil_img.save(resized_buffer, format='JPEG', quality=70, optimize=True)
            resized_buffer.seek(0)

            xl_img = XLImage(resized_buffer)
            xl_img.width = height
            xl_img.height = height
            ws.row_dimensions[row_for_height].height = max(
                ws.row_dimensions[row_for_height].height or 0, height * 0.75
            )
            ws.add_image(xl_img, anchor_cell)
            return True
        except Exception:
            return False

    def _dianke_codigo_anclado(self, product):
        """Códigos de barras alternos/anclados del producto (mismo patrón
        que ya se usa en el Excel de compras a proveedores)."""
        additional = self.env['product.barcode.multi'].search([('product_id', '=', product.id)])
        return ", ".join(additional.mapped('name')) if additional else ''

    @staticmethod
    def _dianke_is_godrej_pocket(display_name):
        """Las 6 referencias de ambientador AMB GODREJ POCKET (Berry Rush,
        Bright, Fresh Blossom, Sea Breeze, Floral Delight, etc.) — Andrés
        las maneja por display, pero Dianke las necesita en unidades
        (1 display = 6 unidades). Se detecta por nombre porque el orden de
        las palabras varía entre referencias (ej. "AMB GODREJ 10GR POCKET
        FLORAL DELIGHT" vs "AMB GODREJ POCKET BERRY RUSH 10GR"), así que se
        exige que tenga GODREJ y POCKET en el nombre, sin importar el
        orden — confirmado por Andrés 2026-09-05."""
        name = (display_name or '').upper()
        return 'GODREJ' in name and 'POCKET' in name

    @staticmethod
    def _dianke_extra_note(display_name, line_name):
        """Devuelve solo la parte de line_name (la descripción/nota que
        escribió el vendedor) que NO es el nombre del producto — ej. si
        line_name es "ALISET NNP 69GR CAMBIO X CAMBIO" y display_name es
        "ALISET NNP 69GR", devuelve "CAMBIO X CAMBIO". Si line_name es
        igual al nombre del producto (o no aporta nada nuevo), devuelve
        cadena vacía en vez de repetir el nombre completo."""
        display_name = (display_name or '').strip()
        line_name = (line_name or '').strip()
        if not line_name:
            return ''
        if not display_name:
            return line_name

        idx = line_name.upper().find(display_name.upper())
        if idx == -1:
            # No hay traslape: el texto es completamente distinto al
            # nombre del producto, se conserva tal cual.
            return line_name

        remainder = line_name[:idx] + line_name[idx + len(display_name):]
        return remainder.strip(' -—.,')

    @staticmethod
    def _dianke_estimate_row_height(text, col_width, base_height=20):
        """Estimación de la altura de fila necesaria para que un texto
        largo con wrap_text no se vea cortado — Excel no la calcula solo
        al generar el archivo por código, hay que aproximarla. ~0.9
        caracteres visibles por unidad de ancho de columna (ajustado
        2026-09-05: la primera estimación de 1.8 se quedaba corta y
        cortaba nombres de producto largos)."""
        if not text:
            return base_height
        chars_per_line = max(int(col_width * 0.9), 8)
        import math
        lineas = math.ceil(len(str(text)) / chars_per_line)
        return max(base_height, 16 * lineas + 8)

    def _fill_dianke_template_sheet(self, ws, rows_data):
        """Hoja única con el formato oficial de pedidos de Dianke (plantilla
        "Formato_para_recibir_pedidos_clientes.xlsx" que Andrés compartió
        2026-09-05), repetido en un bloque por cada orden de rows_data. Se
        agregan datos que la plantilla de Dianke no trae pero Andrés pidió
        de todas formas: RUC, foto del local (en su propia fila, integrada
        al bloque) y Código Anclado en el detalle."""
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        N_COLS = 6  # Código, Código Anclado, Descripción, Cantidad, Precio venta, Tipo de venta
        FONT_NAME = "Aptos"  # misma fuente que usa la plantilla de Dianke
        DESCRIPCION_COL = 3

        LABEL_FONT = Font(name=FONT_NAME, bold=True, size=10, color="173B4D")
        VALUE_FONT = Font(name=FONT_NAME, size=10)
        PAGO_FONT = Font(name=FONT_NAME, bold=True, size=11)
        VALUE_FILL = PatternFill(start_color="FFF9E8", end_color="FFF9E8", fill_type="solid")
        DIVIDER_FILL = PatternFill(start_color="E8F1F5", end_color="E8F1F5", fill_type="solid")
        SECTION_FILL = PatternFill(start_color="173B4D", end_color="173B4D", fill_type="solid")
        SECTION_FONT = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
        TABLE_HEADER_FILL = PatternFill(start_color="2F6F7E", end_color="2F6F7E", fill_type="solid")
        TABLE_HEADER_FONT = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
        ROW_FONT = Font(name=FONT_NAME, size=10)
        ROW_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        NOTA_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        THIN_SIDE = Side(style='thin', color='000000')
        THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

        column_widths = {1: 22, 2: 20, 3: 44, 4: 12, 5: 14, 6: 16}
        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
        descripcion_width = column_widths[DESCRIPCION_COL]

        row_idx = 0
        for data in rows_data:
            order = data['order']
            entrega = data['fecha_entrega']
            entrega_str = entrega.strftime('%d/%m/%Y') if entrega else ''
            efectivo, tarjeta, ach, otro_label = self._dianke_payment_checkboxes(order.custom_payment_method)

            # --- Foto del local (fila propia, integrada al bloque) ---
            if data['partner'] and data['partner'].image_1920:
                row_idx += 1
                foto_label_cell = ws.cell(row=row_idx, column=1, value="Foto del local")
                foto_label_cell.font = LABEL_FONT
                foto_label_cell.border = THIN_BORDER
                ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=N_COLS)
                for col in range(2, N_COLS + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.fill = VALUE_FILL
                    cell.border = THIN_BORDER
                ws.row_dimensions[row_idx].height = 90
                self._dianke_embed_partner_photo(
                    ws, data['partner'], "%s%s" % (get_column_letter(2), row_idx), row_idx, height=110,
                )

            # --- Campos del pedido (etiqueta en A, valor fusionado B:F) ---
            if data['ruta'] and data['orden_ruta']:
                ruta_texto = "%s (Orden %s)" % (data['ruta'], data['orden_ruta'])
            else:
                ruta_texto = data['ruta']
            campos = [
                ("Fecha", data['fecha']),
                ("Nombre o razón social del negocio", data['local']),
                ("RUC", data['ruc']),
                ("Nombre del contacto o persona que recibe el pedido", data['contacto']),
                ("Dirección exacta con indicaciones claras", data['direccion']),
                ("Teléfono de quien recibe el pedido", data['telefono'] or data['celular']),
                ("Número de ruta", ruta_texto),
                ("Número de pedido", order.name),
                ("Fecha en que se debe entregar el pedido al cliente", entrega_str),
            ]
            for label, valor in campos:
                row_idx += 1
                label_cell = ws.cell(row=row_idx, column=1, value=label)
                label_cell.font = LABEL_FONT
                label_cell.border = THIN_BORDER
                ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=N_COLS)
                for col in range(2, N_COLS + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.fill = VALUE_FILL
                    cell.font = VALUE_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_idx, column=2, value=valor)
                ws.row_dimensions[row_idx].height = 24

            # --- Tipo de pago (casillas) ---
            row_idx += 1
            pago_label_cell = ws.cell(row=row_idx, column=1, value="Tipo de pago")
            pago_label_cell.font = LABEL_FONT
            pago_label_cell.border = THIN_BORDER
            opciones_pago = [("Efectivo", efectivo), ("Tarjeta", tarjeta), ("ACH", ach)]
            if otro_label:
                opciones_pago.append((otro_label, True))
            for i in range(N_COLS - 1):
                col = 2 + i
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = VALUE_FILL
                cell.border = THIN_BORDER
                if i < len(opciones_pago):
                    texto, marcado = opciones_pago[i]
                    cell.value = "%s %s" % ("☑" if marcado else "☐", texto)
                    cell.font = PAGO_FONT
                cell.alignment = Alignment(horizontal='center')
            ws.row_dimensions[row_idx].height = 24

            # --- Fila divisoria ---
            row_idx += 1
            for col in range(1, N_COLS + 1):
                ws.cell(row=row_idx, column=col).fill = DIVIDER_FILL if col == 1 else VALUE_FILL
            ws.row_dimensions[row_idx].height = 10

            # --- DETALLE DEL PEDIDO ---
            row_idx += 1
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=N_COLS)
            cell = ws.cell(row=row_idx, column=1, value="DETALLE DEL PEDIDO")
            cell.fill = SECTION_FILL
            cell.font = SECTION_FONT
            cell.alignment = Alignment(horizontal='center')
            ws.row_dimensions[row_idx].height = 22

            # --- Encabezado de la tabla de productos ---
            row_idx += 1
            headers_tabla = ["Código", "Código Anclado", "Descripción", "Cantidad", "Precio venta", "Tipo de venta"]
            for col, texto in enumerate(headers_tabla, start=1):
                cell = ws.cell(row=row_idx, column=col, value=texto)
                cell.fill = TABLE_HEADER_FILL
                cell.font = TABLE_HEADER_FONT
                cell.alignment = Alignment(horizontal='center')
            ws.row_dimensions[row_idx].height = 22

            # --- Líneas de producto ---
            for line in data['lines']:
                row_idx += 1
                product = line.product_id
                codigo = product.barcode or product.default_code or ''
                codigo_anclado = self._dianke_codigo_anclado(product)
                nota = self._dianke_extra_note(product.display_name, line.name)
                tipo_venta = nota if nota else "Normal"
                descripcion = product.display_name or ''

                cantidad = line.product_uom_qty
                precio = line.price_unit
                if self._dianke_is_godrej_pocket(product.display_name):
                    cantidad = cantidad * DIANKE_GODREJ_POCKET_UNITS_PER_DISPLAY
                    precio = precio / DIANKE_GODREJ_POCKET_UNITS_PER_DISPLAY
                precio = round(precio, 2)  # nunca 3+ decimales, pedido de Andrés 2026-09-05

                valores = [codigo, codigo_anclado, descripcion, cantidad, precio, tipo_venta]
                for col, valor in enumerate(valores, start=1):
                    cell = ws.cell(row=row_idx, column=col, value=valor)
                    cell.fill = NOTA_FILL if nota else ROW_FILL
                    cell.font = ROW_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(
                        horizontal='left' if col in (2, 3) else 'center',
                        vertical='center',
                        wrap_text=(col in (2, 3, 6)),
                    )
                    if col == 5:  # Precio venta: formato fijo de 2 decimales, sin importar el cálculo
                        cell.number_format = '0.00'
                ws.row_dimensions[row_idx].height = max(
                    self._dianke_estimate_row_height(codigo_anclado, column_widths[2]),
                    self._dianke_estimate_row_height(descripcion, descripcion_width),
                    self._dianke_estimate_row_height(tipo_venta, column_widths[6]),
                )

            # --- Separación entre pedidos ---
            row_idx += 2

        ws.freeze_panes = None

    def _dianke_subject_and_body(self, orders):
        """Arma asunto y cuerpo del correo a Dianke, en singular o plural
        según la cantidad de órdenes — texto acordado con Andrés
        2026-09-05. El asunto muestra las rutas con la cantidad de
        órdenes de cada una (ej. "Pedregal (5), Villa Grecia (3)") en vez
        de listar cada número de orden, para que no quede kilométrico
        cuando son muchas órdenes — pedido de Andrés 2026-09-05."""
        names = orders.mapped('name')
        names_str = ", ".join(names)
        plural = len(names) > 1

        route_counts = {}
        route_order_list = []
        for order in orders.sorted(key=lambda o: o.name):
            ruta, _ = self._dianke_route_info(order.partner_id)
            key = ruta or 'Sin Ruta'
            if key not in route_counts:
                route_counts[key] = 0
                route_order_list.append(key)
            route_counts[key] += 1
        rutas_str = ", ".join("%s (%d)" % (r, route_counts[r]) for r in route_order_list)
        fecha_str = self._dianke_fecha_larga_es(fields.Date.context_today(self))

        if plural:
            subject = "Pedidos para Dianke Group — %s — %s" % (rutas_str, fecha_str)
            pedido_texto = "los pedidos confirmados"
            listo_texto = "listos"
        else:
            subject = "Pedido para Dianke Group — %s — %s" % (rutas_str, fecha_str)
            pedido_texto = "el pedido confirmado"
            listo_texto = "listo"

        body = """
<div style="margin: 0px; padding: 0px;">
    <p style="margin: 0px; padding: 0px; font-size: 13px;">
        Querido equipo de Dianke,
        <br/><br/>
        Adjunto el Excel con %s (<strong>%s</strong>), %s para su revisión e importación al sistema.
        Ahí está toda la información acordada; si necesitan algún ajuste, por favor identifíquenlo
        para poder corregirlo.
        <br/><br/>
        Así mismo, solicitamos su especial atención a las siguientes especificaciones de despacho:
        <br/><br/>
        <ul style="margin: 0px; padding-left: 20px;">
            <li>Presentación de Productos NNP: todos los artículos se solicitan en unidades individuales.</li>
            <li>Ambientadores Pocket: también se envían en unidades — ya se ajustó la conversión desde display para facilitarles la recepción e importación.</li>
        </ul>
        <br/>
        Quedo a su disposición ante cualquier consulta adicional.
        <br/><br/>
        ¿Podría confirmar que recibió esta orden?
        <br/><br/>
        Atentamente,<br/>
        Andrés Gutiérrez<br/>
        Asistente<br/>
        Shalom Panamá.
        <br/><br/>
    </p>
</div>
""" % (pedido_texto, names_str, listo_texto)

        return subject, body

    def _send_dianke_export_email(self):
        """Envía por correo un XLSX por cada ruta (órdenes de venta
        confirmadas) a Dianke y marca cada orden como exportada."""
        orders = self.filtered(lambda o: o.state == 'sale')
        if not orders:
            return False

        files = orders._generate_dianke_xlsx_files()
        attachment_ids = []
        for filename, xlsx_bytes in files:
            attachment = self.env['ir.attachment'].create({
                'name': filename,
                'type': 'binary',
                'raw': xlsx_bytes,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'res_model': 'sale.order',
                'res_id': orders[0].id,
            })
            attachment_ids.append(attachment.id)

        subject, body = self._dianke_subject_and_body(orders)

        mail = self.env['mail.mail'].create({
            'subject': subject,
            'body_html': body,
            'email_to': DIANKE_EMAIL_TO,
            'email_cc': DIANKE_EMAIL_CC,
            'attachment_ids': [(6, 0, attachment_ids)],
            'model': 'sale.order',
            'res_id': orders[0].id,
        })
        mail.send()

        now = fields.Datetime.now()
        orders.write({
            'custom_dianke_exported': True,
            'custom_dianke_exported_date': now,
            'custom_dianke_exported_by': False,
        })
        for order in orders:
            order.message_post(
                body="Enviado a Dianke automáticamente el %s." % now,
            )
        return True

    def action_send_dianke_export_now(self):
        """Botón/acción manual: genera un Excel por ruta de las órdenes
        seleccionadas que estén confirmadas (ignora las que no estén en
        estado 'sale') y abre la ventana de confirmación para revisar/
        editar antes de enviar."""
        orders = self.filtered(lambda o: o.state == 'sale')
        if not orders:
            from odoo.exceptions import UserError
            raise UserError("Selecciona al menos una orden de venta CONFIRMADA para enviar a Dianke.")

        files = orders._generate_dianke_xlsx_files()
        attachment_ids = []
        for filename, xlsx_bytes in files:
            attachment = self.env['ir.attachment'].create({
                'name': filename,
                'type': 'binary',
                'raw': xlsx_bytes,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'res_model': 'sale.order',
                'res_id': orders[0].id,
            })
            attachment_ids.append(attachment.id)

        subject, body = self._dianke_subject_and_body(orders)

        wizard = self.env['sale.dianke.email.wizard'].create({
            'sale_order_ids': [(6, 0, orders.ids)],
            'email_to': DIANKE_EMAIL_TO,
            'email_cc': DIANKE_EMAIL_CC,
            'subject': subject,
            'body': body,
            'attachment_ids': [(6, 0, attachment_ids)],
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'sale.dianke.email.wizard',
            'res_id': wizard.id,
            'view_mode': 'form',
            'target': 'new',
        }

    @api.model
    def _cron_send_dianke_export(self):
        """Job automático (11:59pm): junta todas las órdenes de venta
        confirmadas que aún no se le han mandado a Dianke y las envía en un
        solo correo con un solo Excel."""
        pending = self.search([
            ('state', '=', 'sale'),
            ('custom_dianke_exported', '=', False),
        ])
        if pending:
            pending._send_dianke_export_email()


class SaleOrderLine(models.Model):
    _inherit = 'sale.order.line'

    custom_product_image = fields.Binary(
        related='product_id.image_1024',
        string='Imagen',
        readonly=True,
    )
    custom_product_barcode = fields.Char(
        related='product_id.barcode',
        string='Código de Barras',
        readonly=True,
    )
    custom_product_default_code = fields.Char(
        related='product_id.default_code',
        string='Código',
        readonly=True,
    )
    custom_promo_status = fields.Char(
        string='Promo',
        compute='_compute_custom_promo_status',
    )

    def _price_qualifies_for_promo(self, line, rule):
        """Verifica si el precio de la línea califica para la promoción."""
        product = line.product_id
        price = line.price_unit

        # Tintes NNP: precio mínimo especial 1.16
        is_tinte_nnp = False
        if rule.product_category_id:
            cat = product.categ_id
            while cat:
                if cat.name and 'tinte' in cat.name.lower() and 'nnp' in cat.name.lower():
                    is_tinte_nnp = True
                    break
                cat = cat.parent_id
        if not is_tinte_nnp and product in rule.product_ids:
            # Verificar por nombre si es tinte NNP
            if 'TINTE NNP' in (product.name or '').upper():
                is_tinte_nnp = True

        if is_tinte_nnp:
            return price >= TINTE_NNP_MIN_PRICE

        # Todos los demás: precio no puede bajar del list_price
        list_price = product.list_price or 0.0
        if list_price <= 0:
            return True
        return price >= list_price - 0.001  # tolerancia de redondeo

    def _get_matched_program(self, line, programs):
        """Retorna (program, rule) que aplica a esta línea, o (None, None)."""
        for program in programs:
            for rule in program.rule_ids:
                if line.product_id in rule.product_ids:
                    return program, rule
                if rule.product_category_id:
                    cat = line.product_id.categ_id
                    while cat:
                        if cat == rule.product_category_id:
                            return program, rule
                        cat = cat.parent_id
        return None, None

    @api.depends(
        'product_id', 'product_uom_qty', 'price_unit',
        'order_id.order_line.product_id',
        'order_id.order_line.product_uom_qty',
        'order_id.order_line.price_unit',
    )
    def _compute_custom_promo_status(self):
        programs = self.env['loyalty.program'].search([
            ('program_type', '=', 'buy_x_get_y'),
            ('active', '=', True),
        ])

        for line in self:
            if not line.product_id or line.display_type:
                line.custom_promo_status = ''
                continue

            program, rule = self._get_matched_program(line, programs)

            if not program or not rule:
                line.custom_promo_status = ''
                continue

            # Verificar precio de esta línea
            if not self._price_qualifies_for_promo(line, rule):
                line.custom_promo_status = ''
                continue

            # Sumar qty de todas las líneas que califican al mismo programa
            total_qty = 0.0
            for ol in line.order_id.order_line:
                if ol.display_type or not ol.product_id:
                    continue
                _, ol_rule = self._get_matched_program(ol, programs)
                if ol_rule != rule:
                    continue
                if not self._price_qualifies_for_promo(ol, rule):
                    continue
                total_qty += ol.product_uom_qty

            min_qty = rule.minimum_qty
            reward = program.reward_ids[:1]
            reward_qty = int(reward.reward_product_qty) if reward else 0
            promo_label = f'{int(min_qty)}+{reward_qty}'

            if min_qty <= 0:
                line.custom_promo_status = ''
                continue

            promos_completas = int(total_qty // min_qty)
            remainder = total_qty % min_qty

            if remainder == 0 and total_qty >= min_qty:
                line.custom_promo_status = f'✅ Tienes {promos_completas} promos completas · Llevas {int(total_qty)} unidades válidas'
            else:
                faltan = int(min_qty - remainder) if remainder > 0 else int(min_qty)
                line.custom_promo_status = f'⏳ Faltan {faltan} unidades · Llevas {int(total_qty)} unidades válidas'
